from __future__ import annotations

from difflib import SequenceMatcher
from functools import lru_cache
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill

from .config import FUZZY_MATCH_AMBIGUITY_MARGIN, FUZZY_MATCH_THRESHOLD
from .utils import clean_text, normalize_key

TOKEN_NORMALIZATION_MAP = {
    # Spanish -> English (common food-service terms)
    "aceite": "oil",
    "aguacate": "avocado",
    "aluminio": "aluminum",
    "amarilla": "yellow",
    "amarillo": "yellow",
    "arroz": "rice",
    "blanca": "white",
    "blanco": "white",
    "blanqueador": "bleach",
    "bolsa": "bag",
    "camaron": "shrimp",
    "camarones": "shrimp",
    "carne": "meat",
    "cebolla": "onion",
    "cerdo": "pork",
    "champinon": "mushroom",
    "chile": "pepper",
    "crudo": "raw",
    "cruda": "raw",
    "desinfectante": "disinfectant",
    "entera": "whole",
    "entero": "whole",
    "espuma": "foam",
    "fresca": "fresh",
    "fresco": "fresh",
    "frijol": "bean",
    "hueso": "bone",
    "huevo": "egg",
    "jalapeno": "jalapeno",
    "leche": "milk",
    "limon": "lime",
    "manteca": "shortening",
    "margarina": "margarine",
    "mezcla": "mix",
    "molida": "ground",
    "muslo": "thigh",
    "papel": "paper",
    "pechuga": "breast",
    "pimiento": "pepper",
    "pollo": "chicken",
    "queso": "cheese",
    "recipiente": "container",
    "res": "beef",
    "rebanada": "sliced",
    "rebanado": "sliced",
    "roja": "red",
    "rojo": "red",
    "sal": "salt",
    "servilleta": "napkin",
    "surtidos": "mixed",
    "tallos": "stem",
    "tapa": "lid",
    "tilapia": "tilapia",
    "tomate": "tomato",
    "tomatillo": "tomatillo",
    "vegetales": "vegetable",
    "verde": "green",
    "vaso": "cup",
    # Common abbreviations seen in listings
    "alum": "aluminum",
    "brst": "breast",
    "chix": "chicken",
    "cmpt": "compartment",
    "comp": "compartment",
    "cont": "container",
    "frsh": "fresh",
    "nugg": "nugget",
    "oz": "ounce",
    "lbs": "pound",
    "lb": "pound",
    "pty": "patty",
    "stk": "steak",
    "wht": "white",
    "ylw": "yellow",
}

GENERIC_TOKENS = {
    "a",
    "an",
    "and",
    "at",
    "by",
    "for",
    "from",
    "in",
    "of",
    "on",
    "or",
    "the",
    "to",
    "with",
    "without",
    "al",
    "con",
    "de",
    "del",
    "el",
    "en",
    "la",
    "las",
    "los",
    "o",
    "para",
    "por",
    "sin",
    "y",
    "anaquel",
    "case",
    "caja",
    "estables",
    "estable",
    "grade",
    "liquida",
    "liquido",
    "pack",
    "ref",
    "seleccion",
    "selecto",
    "suelto",
    "tipo",
    "unidad",
    "unidades",
    "units",
}


@lru_cache(maxsize=8192)
def _product_features(product_name: str) -> tuple[str, frozenset[str], frozenset[str], frozenset[str]]:
    normalized = normalize_key(product_name)
    tokens: list[str] = []

    for raw_token in normalized.split():
        mapped = TOKEN_NORMALIZATION_MAP.get(raw_token, raw_token)
        for token in mapped.split():
            if len(token) > 4 and token.endswith("s"):
                token = token[:-1]
            if token in GENERIC_TOKENS:
                continue
            tokens.append(token)

    canonical_text = " ".join(tokens)
    token_set = frozenset(tokens)
    strong_tokens = frozenset(
        token for token in tokens if len(token) >= 4 and not any(c.isdigit() for c in token)
    )
    number_tokens = frozenset(token for token in tokens if any(c.isdigit() for c in token))
    return canonical_text, token_set, strong_tokens, number_tokens


def _product_similarity(name_a: str, name_b: str) -> float:
    canonical_a, tokens_a, strong_a, numbers_a = _product_features(name_a)
    canonical_b, tokens_b, strong_b, numbers_b = _product_features(name_b)

    if not tokens_a or not tokens_b:
        return 0.0

    shared_numbers = numbers_a & numbers_b
    if numbers_a and numbers_b and not shared_numbers:
        return 0.0

    shared_strong = strong_a & strong_b
    if not shared_strong and not shared_numbers:
        return 0.0

    overlap_count = len(tokens_a & tokens_b)
    jaccard = overlap_count / len(tokens_a | tokens_b)
    containment = overlap_count / min(len(tokens_a), len(tokens_b))
    sequence_ratio = SequenceMatcher(None, canonical_a, canonical_b).ratio()

    score = 0.50 * max(jaccard, containment) + 0.25 * jaccard + 0.25 * sequence_ratio
    if shared_numbers:
        score += 0.08
    if len(shared_strong) >= 2:
        score += 0.05
    return min(score, 1.0)


def _choose_cluster_for_product(
    clusters: list[dict[str, object]],
    store_name: str,
    product_key: str,
    product_name: str,
) -> tuple[int, float, float]:
    best_idx = -1
    best_score = 0.0
    second_score = 0.0

    for idx, cluster in enumerate(clusters):
        prices: dict[str, float] = cluster["prices"]
        if store_name in prices:
            continue

        if str(cluster["product_key"]) == product_key:
            return idx, 1.0, 0.0

        items: list[dict[str, object]] = cluster["items"]
        score = max(
            _product_similarity(product_name, str(item["product_name"]))
            for item in items
        )

        if score > best_score:
            second_score = best_score
            best_score = score
            best_idx = idx
        elif score > second_score:
            second_score = score

    return best_idx, best_score, second_score


def build_comparison_dataframe(frames_by_store: dict[str, pd.DataFrame]) -> pd.DataFrame:
    ordered_store_names = list(frames_by_store.keys())
    if not ordered_store_names:
        raise ValueError("Listings could not be combined.")

    clusters: list[dict[str, object]] = []
    for store_name in ordered_store_names:
        frame = frames_by_store[store_name][["product_key", "product_name", "price"]].copy()
        frame = frame.dropna(subset=["product_name", "price"])
        frame["product_name"] = frame["product_name"].map(clean_text)
        frame = frame[frame["product_name"] != ""].sort_values(
            ["product_name", "price"], kind="stable"
        )

        for row in frame.itertuples(index=False):
            product_name = str(row.product_name)
            product_key = str(row.product_key)
            price = float(row.price)

            cluster_idx, best_score, second_score = _choose_cluster_for_product(
                clusters=clusters,
                store_name=store_name,
                product_key=product_key,
                product_name=product_name,
            )

            should_merge = (
                cluster_idx >= 0
                and best_score >= FUZZY_MATCH_THRESHOLD
                and (best_score - second_score) >= FUZZY_MATCH_AMBIGUITY_MARGIN
            )

            if not should_merge:
                clusters.append(
                    {
                        "product_key": product_key,
                        "product_name": product_name,
                        "prices": {store_name: price},
                        "items": [
                            {
                                "store_name": store_name,
                                "product_name": product_name,
                            }
                        ],
                    }
                )
                continue

            cluster = clusters[cluster_idx]
            cluster["items"].append(
                {
                    "store_name": store_name,
                    "product_name": product_name,
                }
            )
            cluster_prices: dict[str, float] = cluster["prices"]
            current_price = cluster_prices.get(store_name)
            if current_price is None or price < current_price:
                cluster_prices[store_name] = price

            current_name = str(cluster["product_name"])
            if len(product_name) < len(current_name):
                cluster["product_name"] = product_name
                cluster["product_key"] = normalize_key(product_name)

    if not clusters:
        raise ValueError("Listings could not be combined.")

    merged_rows: list[dict[str, object]] = []
    for cluster in clusters:
        row: dict[str, object] = {
            "product_key": str(cluster["product_key"]),
            "Product": str(cluster["product_name"]),
        }
        cluster_prices: dict[str, float] = cluster["prices"]
        for store_name in ordered_store_names:
            row[store_name] = cluster_prices.get(store_name)
        merged_rows.append(row)

    merged_df = pd.DataFrame(merged_rows)
    merged_df = merged_df[["product_key", "Product"] + ordered_store_names]
    merged_df = merged_df.sort_values("Product", kind="stable").reset_index(drop=True)
    return merged_df


def build_ranking_dataframe(
    comparison_df: pd.DataFrame, store_columns: list[str]
) -> pd.DataFrame:
    best_price_counts = {store: 0 for store in store_columns}

    for _, row in comparison_df.iterrows():
        available_prices: dict[str, float] = {}
        for store in store_columns:
            value = row[store]
            if pd.isna(value):
                continue
            price = float(value)
            if price <= 0:
                continue
            available_prices[store] = price

        if not available_prices:
            continue

        min_price = min(available_prices.values())
        for store, price in available_prices.items():
            if abs(price - min_price) < 1e-9:
                best_price_counts[store] += 1

    ranking_df = pd.DataFrame(
        [
            {
                "Store": store,
                "Products with best price": best_price_counts[store],
            }
            for store in store_columns
        ]
    )

    ranking_df = ranking_df.sort_values(
        ["Products with best price", "Store"],
        ascending=[False, True],
    ).reset_index(drop=True)
    return ranking_df


def rgb_gradient_green_white(ratio: float) -> str:
    ratio = max(0.0, min(1.0, ratio))
    # Vivid green for stronger visual contrast while keeping green -> white scale.
    green = (0, 200, 0)
    white = (255, 255, 255)
    red = round(green[0] + (white[0] - green[0]) * ratio)
    g_value = round(green[1] + (white[1] - green[1]) * ratio)
    blue = round(green[2] + (white[2] - green[2]) * ratio)
    return f"FF{red:02X}{g_value:02X}{blue:02X}"


def apply_price_gradient(ws, price_col_idx: list[int]) -> None:
    gradient_steps = 11
    fill_by_step = {}
    for step in range(gradient_steps):
        ratio = step / (gradient_steps - 1)
        color = rgb_gradient_green_white(ratio)
        fill_by_step[step] = PatternFill(
            fill_type="solid",
            start_color=color,
            end_color=color,
        )

    for row_idx in range(2, ws.max_row + 1):
        numeric_values: list[float] = []
        parsed_row: dict[int, float] = {}

        for col_idx in price_col_idx:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None or cell.value == "":
                continue
            try:
                value = float(cell.value)
            except (TypeError, ValueError):
                continue
            if value <= 0:
                continue
            parsed_row[col_idx] = value
            numeric_values.append(value)

        if not numeric_values:
            continue

        row_min = min(numeric_values)
        row_max = max(numeric_values)

        for col_idx, value in parsed_row.items():
            ratio = 0.0 if row_max == row_min else (value - row_min) / (row_max - row_min)
            # Quantize to visible bands so intermediate differences stand out.
            step = round(ratio * (gradient_steps - 1))
            ws.cell(row=row_idx, column=col_idx).fill = fill_by_step[step]
            ws.cell(row=row_idx, column=col_idx).number_format = "$#,##0.00"


def autosize_columns(ws, max_width: int = 45) -> None:
    for column_cells in ws.columns:
        length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), max_width)


def export_report(
    output_path: Path,
    comparison_df: pd.DataFrame,
    ranking_df: pd.DataFrame,
    store_columns: list[str],
    relations_df: pd.DataFrame | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        export_df = comparison_df.drop(columns=["product_key"])
        export_df.to_excel(writer, index=False, sheet_name="Comparison")
        ranking_df.to_excel(writer, index=False, sheet_name="Ranking")
        if relations_df is not None and not relations_df.empty:
            relations_df.to_excel(writer, index=False, sheet_name="Relations")

        ws_comp = writer.sheets["Comparison"]
        ws_rank = writer.sheets["Ranking"]

        ws_comp.freeze_panes = "B2"
        price_col_idx = list(range(2, 2 + len(store_columns)))
        apply_price_gradient(ws_comp, price_col_idx)
        autosize_columns(ws_comp)
        autosize_columns(ws_rank)

        if "Relations" in writer.sheets:
            ws_rel = writer.sheets["Relations"]
            autosize_columns(ws_rel)
